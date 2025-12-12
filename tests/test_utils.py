import datetime
import textwrap

import openpyxl
import pyquoks

import schedule_parser


class TestUtils(pyquoks.test.TestCase):
    _MODULE_NAME = __name__

    @classmethod
    def setUpClass(cls) -> None:
        super().setUpClass()

        cls._substitutions_names = [
            "substitutions",
            "substitutions_old",
            "substitutions_incorrect",
        ]

        cls._group = "СИСАД 24-01"
        cls._schedule_with_substitutions = {
            "08_11_25": "",
            "18_11_25": textwrap.dedent(
                """\
                0. Разг. о важном | 45 к.
                1. ААС | 34 к.
                2. Операц. СиС | 34 к.
                3. ТФУПД | 44 к.
                4. Иностр. яз. | 1 к.
                """,
            ),
            "25_11_25": textwrap.dedent(
                """\
                0. Разг. о важном | 45 к.
                1. (1) Иностр. яз. | 1 к.
                1. (2) Иностр. яз. | 23 к.
                2. Комп. сети (П) | 45 к.
                3. ААС | 34 к.
                4. Операц. СиС | 34 к.
                """,
            ),
            "28_11_25": textwrap.dedent(
                """\
                1. Кл. час | 45 к.
                2. Комп. сети (Л) | 45 к.
                3. Комп. сети (П) | 45 к.
                4. Элементы ВМ | 40 к.
                5. Элементы ВМ | э/о
                """,
            ),
            "29_11_25": "",
            "02_12_25": textwrap.dedent(
                """\
                1. Разг. о важном | 45 к.
                2. Комп. сети (Л) | 33 к.
                3. ТФУПД | 11 к.
                4. (1) Иностр. яз. | 1 к.
                4. (2) Иностр. яз. | 23 к.
                """,
            ),
            "06_12_25": textwrap.dedent(
                """\
                0. Кл. час | 45 к.
                1. Основы проектирования баз данных | 34 к.
                2. Основы проектирования баз данных | 34 к.
                3. ТФУПД | 44 к.
                """,
            ),
        }

    def test_parse_schedule(self):
        workbook = openpyxl.load_workbook(
            filename=pyquoks.utils.get_path("resources/tables/schedule.xlsx"),
        )

        for group in list(schedule_parser.utils.parse_schedule(workbook.worksheets[0])):
            self.assert_type(
                func_name=self.test_parse_schedule.__name__,
                test_data=group,
                test_type=schedule_parser.models.GroupSchedule,
                message="objects in parsed schedules list",
            )

    def test_parse_substitutions(self):
        for name in self._substitutions_names:
            workbook = openpyxl.load_workbook(
                filename=pyquoks.utils.get_path(f"resources/tables/{name}.xlsx"),
            )

            for substitution in list(schedule_parser.utils.parse_substitutions(workbook.worksheets[0])):
                self.assert_type(
                    func_name=self.test_parse_substitutions.__name__,
                    test_data=substitution,
                    test_type=schedule_parser.models.Substitution,
                    message=f"({name}) objects in parsed substitutions list",
                )

    def test_schedule_with_substitutions(self):
        workbook_schedule = openpyxl.load_workbook(
            filename=pyquoks.utils.get_path("resources/tables/schedule.xlsx"),
        )

        for name, correct_schedule in self._schedule_with_substitutions.items():
            workbook_substitutions = openpyxl.load_workbook(
                filename=pyquoks.utils.get_path(f"resources/tables/substitutions_{name}.xlsx"),
            )

            current_date = datetime.datetime.strptime(name, "%d_%m_%y")

            current_schedule = schedule_parser.utils.get_schedule_with_substitutions(
                schedule=list(
                    schedule_parser.utils.parse_schedule(
                        worksheet=workbook_schedule.worksheets[0],
                    )
                ),
                substitutions=list(
                    schedule_parser.utils.parse_substitutions(
                        worksheet=workbook_substitutions.worksheets[0],
                    )
                ),
                group=self._group,
                date=current_date,
            )

            self.assert_equal(
                func_name=self.test_schedule_with_substitutions.__name__,
                test_data="".join(f"{period.readable}\n" for period in current_schedule),
                test_expected=correct_schedule,
                message=f"({name}) compare parsed schedule with correct one",
            )
